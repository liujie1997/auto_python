from openpyxl import *


class Excel():

    def read_excel(self, filename, sheet_name):
        wb = load_workbook(filename)
        sh = wb[sheet_name]

    def write_excel(self, filename, sheet_name, data, header):
        try:
            wb = load_workbook(filename)
            ws = wb[sheet_name]
        except:
            wb = Workbook()
            ws = wb.create_sheet(sheet_name, 0)
            ws.merge_cells(range_string="A1:B1")
            ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=1)
            ws.cell(1, 1, header)
        for i in range(0, len(data)):
            ws.cell(i + 2, 1).value = list(data.keys())[i]
            ws.cell(i + 2, 2).value = list(data.values())[i]
        wb.save(filename)
